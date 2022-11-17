
from openpyxl import Workbook, load_workbook
import softest
class Utils():
    #softest.TestCase
      def __init__(self,browser):
         self.browser= browser

      def assert_list_items(self,list, value):

          for stop in list:
              assert (stop.text) == value
              print("assert passed")

      def scroll_to_elemInList_AndClick(self,list):

          for elem in list:

              self.browser.execute_script("arguments[0].scrollIntoView();", elem)
              self.browser.execute_script("arguments[0].click();", elem)


      def read_data_from_excel(file_name,sheet):
          print("hello")
          datalist = []
          wb = load_workbook(filename=file_name)
          sh = wb[sheet]
          row_ct = sh.max_row
          col_ct = sh.max_column

          for i in range(2, row_ct + 1):
              row = []
              for j in range(1, col_ct + 1):
                  row.append(sh.cell(row=i, column=j).value)
              datalist.append(row)

          return datalist
